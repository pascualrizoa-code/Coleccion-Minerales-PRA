import openpyxl
import json
import datetime
import shutil
import os

EXCEL_FILE = "Coleccion de minerales.xlsx"
JSON_FILE = "catalogo_minerales.json"

# Columnas del JSON y sus equivalentes en el Excel
# El orden debe coincidir con las cabeceras del Excel
COLUMN_MAP = {
    "Nº Inventario": "Nº Inventario",
    "Mineral": "Mineral",
    "Variedad": "Variedad",
    "Min_asociado": "Min_asociado",
    "Fórmula química": "Fórmula química",
    "Sistema cristalino": "Sistema cristalino",
    "Clase química": "Clase química",
    "Hábito / Morfología": "Hábito / Morfología",
    "Color": "Color",
    "Brillo": "Brillo",
    "Transparencia": "Transparencia",
    "Fluorescencia UVA": "Fluorescencia UVA",
    "Dimensiones (mm)": "Dimensiones (mm)",
    "Cristal (mm)": "Cristal (mm)",
    "Peso (Gramos)": "Peso (Gramos)",
    "Yacimiento": "Yacimiento",
    "Pais": "Pais",
    "Estado de conservación": "Estado de conservación",
    "Intervención conservación": "Intervención conservación",
    "Fecha Adquisición": "Fecha Adquisición",
    "Precio Compra": "Precio Compra",
    "Notas": "Notas",
    "Info": "Info",
    "Valor estimado (€)": "Valor estimado (€)",
    "Fecha Tasación": "Fecha Tasación",
    "Tasador": "Tasador",
}

# Campos numéricos
NUMERIC_FIELDS = {"Peso (Gramos)", "Precio Compra", "Valor estimado (€)"}

# Campos de fecha
DATE_FIELDS = {"Fecha Adquisición", "Fecha Tasación"}

def clean_value(value, field_name):
    """Limpia y convierte el valor según el tipo de campo."""
    if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
        return None
    
    # Fechas
    if field_name in DATE_FIELDS:
        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, str):
            return value.strip()
        return None
    
    # Numéricos
    if field_name in NUMERIC_FIELDS:
        try:
            if isinstance(value, float) and value == int(value):
                return int(value)
            return value
        except (ValueError, TypeError):
            return None
    
    # Strings
    if isinstance(value, str):
        cleaned = value.strip()
        # Eliminar _x000D_ (retorno de carro de Excel)
        cleaned = cleaned.replace("_x000D_", "").replace("\r\n", "\n").replace("\r", "\n")
        return cleaned if cleaned else None
    
    if isinstance(value, float) and value == int(value):
        return int(value)
    
    return value


def excel_to_json():
    print(f"Leyendo {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Leer cabeceras de la primera fila
    headers = [cell.value for cell in ws[1]]
    print(f"Cabeceras encontradas: {headers}")
    print(f"Total de filas de datos: {ws.max_row - 1}")

    # Hacer backup del JSON actual
    if os.path.exists(JSON_FILE):
        from datetime import datetime as dt
        backup_name = f"catalogo_minerales_backup_{dt.now().strftime('%Y%m%d_%H%M%S')}.json"
        shutil.copy2(JSON_FILE, backup_name)
        print(f"Backup creado: {backup_name}")

    # Construir lista de registros
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas completamente vacías
        if all(v is None or v == "" for v in row):
            continue

        record = {}
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            cleaned = clean_value(value, header)
            record[header] = cleaned

        # Solo incluir si tiene número de inventario
        inv = record.get(headers[0])
        if inv:
            records.append(record)

    print(f"Registros procesados: {len(records)}")

    # Guardar JSON
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"JSON guardado en {JSON_FILE}")
    
    # Mostrar M-004 para verificar
    for r in records:
        if r.get(headers[0]) == "M-004":
            print("\n--- M-004 ---")
            try:
                print(json.dumps(r, ensure_ascii=False, indent=2))
            except UnicodeEncodeError:
                print(json.dumps(r, ensure_ascii=True, indent=2))
            break

    return len(records)


if __name__ == "__main__":
    count = excel_to_json()
    print(f"\nTotal minerales en el catálogo: {count}")
